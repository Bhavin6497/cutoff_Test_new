import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import lasio
import numpy as np
import os
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import xlwings as xw
import plotly.io as pio
import streamlit as st
import io
from st_aggrid import AgGrid, GridOptionsBuilder,GridUpdateMode
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
import tempfile
st.set_page_config(layout="wide")

# Center-align the title
st.markdown(
    """
    <style>
    .title {
        text-align: center;
    }
    </style>
    """,
    unsafe_allow_html=True
)

st.markdown("<h1 class='title'>CUTOFF - ANALYSER</h1>", unsafe_allow_html=True)
# Folder path where LAS files are located
st.sidebar.title('APP_CHRONOLOGY')
def get_session_state():
    if 'files' not in st.session_state:
        st.session_state.files= None
    if 'session_state.df_final' not in st.session_state:
        st.session_state.df_final= None
    if 'No_of_rows' not in st.session_state:
        st.session_state.No_of_rows= None
    if 'NTG' not in st.session_state:
        st.session_state.NTG= None
    if 'PVH' not in st.session_state:
        st.session_state.PVH= None
    if 'Vclay_Cutoff_df' not in st.session_state:
        st.session_state['Vclay_Cutoff_df']= None
    if 'Porosity_Cutoff_df' not in st.session_state:
        st.session_state['Porosity_Cutoff_df']= None
    if 'Top_Bot' not in st.session_state:
        st.session_state.Top_Bot= {}
    if 'error_dfs' not in st.session_state:
        st.session_state.error_dfs= {}
    if 'dict_wells' not in st.session_state:
        st.session_state.dict_wells= {}
    if 'dict_Rawdataframe' not in st.session_state:
        st.session_state.dict_Rawdataframe= {}
    if 'Ex_File' not in st.session_state:
        st.session_state.Ex_File= None
    if 'dfs' not in st.session_state:
        st.session_state.dfs = {}
    if 'updated_df' not in st.session_state:
        st.session_state.updated_df = None
    if 'cases' not in st.session_state:
        st.session_state.cases = {}
    if 'df_final_case' not in st.session_state:
        st.session_state.df_final_case = None
    if 'dict1' not in st.session_state:
        st.session_state.dict1 = {}
    if 'dict_Pay_P' not in st.session_state:
        st.session_state.dict_Pay_P = {}
    if 'interval_' not in st.session_state: 
        st.session_state.interval_ = []
    if 'dict2' not in st.session_state:
        st.session_state.dict2 = {}
    if 'dict3' not in st.session_state:
        st.session_state.dict3 = {}
    if 'dict4' not in st.session_state:
        st.session_state.dict4 = {}
    if 'dict5' not in st.session_state:
        st.session_state.dict5 = {}
    if 'dict_Pay_Parameter' not in st.session_state:
        st.session_state.dict_Pay_Parameter = {}
    if 'vclay_cutoff_input' not in st.session_state:
        st.session_state.vclay_cutoff_input=0.6
    if 'pige_cutoff_input' not in st.session_state:
        st.session_state.pige_cutoff_input=0.06
    if 'sw_cutoff_input' not in st.session_state:
        st.session_state.sw_cutoff_input=0.75
    if 'show_plots' not in st.session_state:
        st.session_state.show_plots = False
    if 'Poro' not in st.session_state:
        st.session_state.Poro="None"
    if 'opti_' not in st.session_state:
        st.session_state.opti_ = {}
    if 'selected_well' not in st.session_state:
        st.session_state.selected_well="None"
    return st.session_state
def Histogram(dictionary_any):
        selected_option = st.selectbox('Choose the well', list(dictionary_any.keys()))
        col1, col2  = st.columns((2, 2))
        with col1:    
            df = dictionary_any[selected_option]
            st.write('Data:', df)
        with col2:
            de = df.describe()
            st.write('Data_Statistics:',de)
        if st.button('Display Histograms'):    
            columns_to_plot = df.columns[1:4]
            for column in columns_to_plot:
                df[column] = pd.to_numeric(df[column], errors='coerce')
            df = df.replace([np.inf, -np.inf], np.nan).dropna(subset=columns_to_plot)
            fig, axs = plt.subplots(1, 3, figsize=(15, 5))  
            for ax, column in zip(axs, columns_to_plot):
                ax.hist(df[column], bins=10, edgecolor='black')
                ax.set_title(f'Histogram of {column}')
                ax.set_xlabel(column)
                ax.set_ylabel('Frequency')
            st.pyplot(fig)

def convert_dfs_to_excel(dfs):
    session_state = get_session_state()
    output= io.BytesIO()    
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        for key, value in dfs.items():   
            value[0].to_excel(writer, sheet_name=key, index=False)
            value[1].to_excel(writer, sheet_name=key, startcol=dfs[key][0].shape[1] + 2, index=False)
        session_state.NTG.to_excel(writer, sheet_name='NTG_summary', index=False)
        session_state.PVH.to_excel(writer, sheet_name='PVH_summary', index=False)
    output.seek(0)  
    wb = load_workbook(output)
    ws1 = wb['NTG_summary']
    chart1 = ScatterChart()
    chart1.title = "NTG"
    chart1.x_axis.title = "Cutoff"
    chart1.y_axis.title = "NTG"
    x_data1 = Reference(ws1, min_col=1, min_row=2, max_row=ws1.max_row, max_col=1)
    for i in range(2, ws1.max_column + 1):
        y_data1 = Reference(ws1, min_col=i, min_row=2, max_row=ws1.max_row)
        series_name1 = ws1.cell(row=1, column=i).value
        series1 = Series(y_data1, x_data1, title_from_data=False, title=series_name1)
        series1.marker.symbol = "circle"
        chart1.series.append(series1)
    chart1.width = 20
    chart1.height = 16

    # Adjust x-axis and y-axis properties
    chart1.x_axis.majorTickMark = 'out'
    chart1.x_axis.scaling.min = 0  # Set minimum value for x-axis
    chart1.x_axis.scaling.max = 1  # Set maximum value for x-axis
    chart1.x_axis.tickLblPos = 'low'  # Set position of tick labels
    chart1.x_axis.majorUnit = 0.05

    chart1.y_axis.majorTickMark = 'out'
    chart1.y_axis.scaling.min = 0  # Set minimum value for y-axis
    chart1.y_axis.scaling.max = 1  # Set maximum value for y-axis
    chart1.y_axis.tickLblPos = 'low'  # Set position of tick labels

    # Create a new chart sheet
    chart_sheet = wb.create_chartsheet(title="NTG Chart")

    # Add chart2 to the chart sheet
    chart_sheet.add_chart(chart1)  #

    # Process second sheet
    ws2 = wb['PVH_summary']
    chart2 = ScatterChart()
    chart2.title = "PVH"
    chart2.x_axis.title = "Cutoff"
    chart2.y_axis.title = "PVH"
    x_data2 = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row, max_col=1)
    for i in range(2, ws2.max_column + 1):
        y_data2 = Reference(ws2, min_col=i, min_row=2, max_row=ws2.max_row)
        series_name2 = ws2.cell(row=1, column=i).value
        series2 = Series(y_data2, x_data2, title_from_data=False, title=series_name2)
        series2.marker.symbol = "circle"
        chart2.series.append(series2)
    chart2.width = 20
    chart2.height = 16            

    # Adjust x-axis and y-axis properties
    chart2.x_axis.majorTickMark = 'out'
    chart2.x_axis.scaling.min = 0  # Set minimum value for x-axis
    chart2.x_axis.scaling.max = 1  # Set maximum value for x-axis
    chart2.x_axis.tickLblPos = 'low'  # Set position of tick labels
    chart2.x_axis.majorUnit = 0.05

    chart2.y_axis.majorTickMark = 'out'
    chart2.y_axis.scaling.min = 0  # Set minimum value for y-axis
    chart2.y_axis.scaling.max = 1  # Set maximum value for y-axis
    chart2.y_axis.tickLblPos = 'low'  # Set position of tick labels


    # Create a new chart sheet
    chart_sheet = wb.create_chartsheet(title="PVH Chart")

    # Add chart2 to the chart sheet
    chart_sheet.add_chart(chart2)  #
    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    
    return final_output.getvalue()

def cutoff_entry_form(Property,Defaultvalue,dictany):
    session_state = get_session_state()
    prefilled_values = dictany.keys()
    default_value = st.text_input(f"Enter the {Property}_Cutoff", Defaultvalue)
    default_value = float(default_value)
    data = {
            "Well": prefilled_values,
            f"{Property}_Cutoff": [default_value] * len(prefilled_values)
            }
    df = pd.DataFrame(data)


    # Create a GridOptionsBuilder object
    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_default_column(editable=True)  # Make all columns editable

    # Configure specific columns
    gb.configure_column("Well", editable=False)  # Make prefilled column non-editable
    gb.configure_column(f"{Property}_Cutoff", editable=True)  # Make user input column editable

    # Use the AgGrid component to display the grid
    grid_options = gb.build()
    grid_response = AgGrid(df, gridOptions=grid_options, editable=True)

    # Get the updated DataFrame after user edits
    updated_df = grid_response['data']
    session_state[f"{Property}_Cutoff_df"] = updated_df
    if st.button("Submit"):
        # Save the updated DataFrame to session state
        st.success(f"{Property} Cutoff values have been submitted.")
    
    return updated_df
            
def page0():
    session_state = get_session_state()
    st.success('LAS FILE UPLOADER')
    tab1, tab2 = st.tabs(["MAIN_TAB", "MANUAL_UPLOAD"])
    with tab1:
        # File uploader for .csv or .las files, allowing multiple file uploads
        session_state.files = st.file_uploader("Choose files", type=["las"], accept_multiple_files=True)
        
        default_col_titles = ['WELL', 'SAND_TOP', 'GOC', 'OWC','SAND_BOTTOM','PERF_TOP','PERF_BOTTOM','R_TYPE']
        col_titles = st.text_input(" ", ", ".join(default_col_titles))
        if col_titles:
            col_titles = col_titles.split(",")
        data = st.text_area(" ")
        if st.button("Submit"):
            if data:
                try:
                    session_state.Top_Bot = pd.read_csv(io.StringIO(data), sep="\t", header=None)
                    if col_titles:
                        session_state.Top_Bot.columns = col_titles
                    st.write("Here is your data:")
                    st.dataframe(session_state.Top_Bot)
                    session_state.Top_Bot['WELL'].sort_values()
                    session_state.dict_wells = session_state.Top_Bot.set_index('WELL').T.to_dict('list')
                except Exception as e:
                    st.error(f"Error reading data: {e}")
            else:
                st.warning("Please paste some data into the text area.")
            
        wells = list(session_state.dict_wells.keys())
        if st.button("Autodetect"):
            if session_state.files is not None:
                # Sort the list of files alphabetically
                files = sorted(session_state.files, key=lambda x: x.name)
                
                porosity_options = ['PIGE_QEPP', 'PIGE', 'PHIE_F', 'PHIE', 'PIGE_F']
                saturation_options = ['SUWI', 'SW_F', 'SW', 'SUWI_QEPP', 'sw', 'Sw_F']
                vcl_options = ['VCL', 'VCL_F', 'VCL_QEPP', 'VCL_GEO_QEPP']
                depth_options = ['DEPT', 'MD', 'DEPTH']
                
                session_state.dfs = {}  # Initialize the dfs dictionary in session_state
                session_state.error_dfs = {}
                # Iterate through each file in the folder
                for i,uploaded_file in enumerate(files):
                    bytes_data = uploaded_file.read()
                    str_io = io.StringIO(bytes_data.decode('Windows-1252'))
                    las = lasio.read(str_io)
                    # Convert LAS data to DataFrame
                    df1 = las.df()
                    df1.reset_index(inplace=True)  
                    
                    # Determine the result columns
                    Porosity_result = next((option for option in porosity_options if option in df1.columns), None)
                    Saturation_result = next((option for option in saturation_options if option in df1.columns), None)
                    Vcl_result = next((option for option in vcl_options if option in df1.columns), None)
                    Depth_result = next((option for option in depth_options if option in df1.columns), None)
                    st.write(Porosity_result)
                    if (Porosity_result is None or Saturation_result is None or Vcl_result is None or Depth_result is None):
                        list_df1=[]
                        list_df1.append(df1.columns)
                        list_df1.append(df1)
                        session_state.error_dfs[wells[i]] = list_df1
                    else:
                        # Rename columns and handle missing columns
                        rename_dict = {Depth_result: 'MD', Porosity_result: 'Pi', Saturation_result: 'Sw', Vcl_result: 'Vcl'}
                        df1.rename(columns=rename_dict, inplace=True)
                        
                        # Ensure only available columns are selected
                        available_columns = ['MD', 'Pi', 'Sw', 'Vcl']
                        
                        df1 = df1[available_columns]
                        
                        # Store the DataFrame in session state
                        session_state.dfs[wells[i]] = df1


                if session_state.error_dfs:
                    st.write("ONE OR MORE LAS FILES COULD NOT BE DETECTED, UPLOAD MANUALLY")
                else:
                    st.write("All LAS FILES SUCCESSFULLY DETECTED")

    with tab2:        
        if session_state.error_dfs:
            session_state.selected_well = st.selectbox("Choose the option", list(session_state.error_dfs.keys()))
            if session_state.selected_well:  
                options_ = st.multiselect('Select the appropriate name for the curves in the order - Depth, Effective Porosity, Water Saturation, Vclay_Content', session_state.error_dfs[session_state.selected_well][0].tolist())
                        
                if options_ and len(options_) == 4:
                    session_state.opti_[session_state.selected_well] = options_   
                            
                    rename_dict = {session_state.opti_[session_state.selected_well][0]: 'MD', session_state.opti_[session_state.selected_well][1]: 'Pi', session_state.opti_[session_state.selected_well][2]: 'Sw', session_state.opti_[session_state.selected_well][3]: 'Vcl'}
                    session_state.error_dfs[session_state.selected_well][1].rename(columns=rename_dict, inplace=True)
                                    
                    # Ensure only available columns are selected
                    available_columns = ['MD', 'Pi', 'Sw', 'Vcl']
                                    
                    session_state.error_dfs[session_state.selected_well][1] = session_state.error_dfs[session_state.selected_well][1][available_columns]
                                    
                    # Store the DataFrame in session state
                    session_state.dfs[session_state.selected_well] = session_state.error_dfs[session_state.selected_well][1]
                    session_state.dfs = {key: session_state.dfs[key] for key in sorted(session_state.dfs)}
                    
    
def page01():       
    # Initialize Dev_data dictionary
    session_state = get_session_state()
    Dev_data = {}

    def create_empty_df(rows, cols):
        return pd.DataFrame('', index=range(rows), columns=cols)

    # Create empty dataframes for each well
    for key in session_state.dict_wells.keys():
        Dev_data[key] = create_empty_df(50, ['MD', 'DEVIATION_ANGLE', 'TVD'])

    # Initialize session state variables if they don't exist
    if 'form_data' not in st.session_state:
        st.session_state.form_data = Dev_data

    if 'current_option' not in st.session_state:
        st.session_state.current_option = None

    # Select box for form options
    option = st.selectbox(
        "Select the appropriate Well",
        session_state.dict_wells.keys(),
        key="form_option"
    )

    # Update the current option and data
    if option != st.session_state.current_option:
        st.session_state.current_option = option
        st.session_state.data = st.session_state.form_data[option]

    # Configure the Ag-Grid options
    gb = GridOptionsBuilder.from_dataframe(st.session_state.data)
    gb.configure_default_column(editable=True)
    gb.configure_grid_options(enableRangeSelection=True)
    grid_options = gb.build()

    # Display the Ag-Grid
    grid_response = AgGrid(st.session_state.data, gridOptions=grid_options, update_mode=GridUpdateMode.VALUE_CHANGED)

    # Update the session state data with the edited grid data
    st.session_state.data = grid_response['data']

    # Submit button
    if st.button('Submit'):
        # Convert the data to a DataFrame and remove rows where all values are None or empty
        df = pd.DataFrame(st.session_state.data)
        df = df.dropna(how='all').replace('', pd.NA).dropna(how='all')
        st.session_state.form_data[option] = df
        st.success(f"Data for {option} submitted successfully!")

    # Iterate over the dict_wells and update with form_data
    index = 7
    for key in session_state.dict_wells.keys():
        if index < len(session_state.dict_wells[key]):
            session_state.dict_wells[key][index] = st.session_state.form_data[key]
        else:
            session_state.dict_wells[key].append(st.session_state.form_data[key])

    # Perform data cleaning and type conversion
    for key, value in session_state.dict_wells.items():
        if value[index].dropna(how='all').replace('', pd.NA).dropna(how='all').empty:
            value[index] = value[index].dropna(how='all').replace('', pd.NA).dropna(how='all')
            value[index] = value[index].astype(float)
        else:
            value[index] = value[index].astype(float)
    st.write(session_state.dict_wells)
    
        
          
        
def page1():
    session_state = get_session_state()
    st.success('DISPLAY THE RAW DATA')
    tab1, tab2 = st.tabs(["RAW_DATA", "PROPETIES_PLOTS"])
    with tab1:
        session_state.dict_Rawdataframe = {}
        for (Well, markers), df in zip(session_state.dict_wells.items(), session_state.dfs.values()):
            Name = Well
            target_value_= markers[0]
            target_value= markers[3]
            min_ = df['MD'].iloc[(df['MD'] - target_value_).abs().idxmin()]
            max_ = df['MD'].iloc[(df['MD'] - target_value).abs().idxmin()]
            df = df.dropna().copy()
            condition = (df['MD'] > min_) & (df['MD'] < max_)
            df = df[condition]
            df = df.sort_values(by='MD')
            session_state.dict_Rawdataframe[Name] = df
        Histogram(session_state.dict_Rawdataframe)
        def convert_dfs_to_excel(dfs):
            output= io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            # Iterate over the data and add traces to the subplots
                for key, value in dfs.items():   
                    value.to_excel(writer, sheet_name=key, index=False)
            processed_data = output.getvalue()
            return processed_data
        if st.button('Save Raw Data to Excel'):
            excel_data = convert_dfs_to_excel(session_state.dict_Rawdataframe)
            st.download_button(label = "Download Excel File",data = excel_data,file_name = "Raw_Data.xlsx",mime= 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    with tab2:
        df_plotted = st.selectbox('Choos the well', list(session_state.dict_Rawdataframe.keys()))
        df_p = session_state.dict_Rawdataframe[df_plotted]
        # Constant difference to check
        # Constant difference to check
        constant_diff = 0.1524

        # Initial x-value for the vertical line
        x_value = 1  # You can change this as needed

        # Create a list to store the segments
        segments = []

        # Convert MD column to a numpy array
        y_values = df_p['MD'].values

        # Loop through y-values to check differences
        for i in range(len(y_values) - 1):
            if abs(y_values[i + 1] - y_values[i] - constant_diff) < 1e-10:  # Adding a tolerance for floating point comparisons
                # Append the vertical line segment to the list if the difference is constant_diff
                segments.append(go.Scatter(x=[x_value, x_value], y=[y_values[i], y_values[i + 1]], mode='lines', line=dict(color='blue', width=4)))

        # Create the figure with subplots
        fig3 = make_subplots(rows=1, cols=4, shared_yaxes=True, column_widths=[0.25, 0.25, 0.25, 0.25])

        # Add scatter plots for each column
        fig3.add_trace(go.Scatter(x=df_p['Vcl'], y=df_p['MD'], mode='lines', name='Vclay'), row=1, col=1)
        fig3.add_trace(go.Scatter(x=df_p['Pi'], y=df_p['MD'], mode='lines', name='Porosity', fill='tozeroy', fillcolor='green'), row=1, col=2)
        fig3.add_trace(go.Scatter(x=df_p['Sw'], y=df_p['MD'], mode='lines', name='Saturation', fill='tozeroy'), row=1, col=3)

        # Add the vertical line segments as the fourth subplot
        for segment in segments:
            fig3.add_trace(segment, row=1, col=4)

        # Update layout
        fig3.update_layout(title_text="CUTOFF_PROPERTIES", height=800, width=1200, shapes=[
            dict(type="rect", xref="paper", yref="paper", x0=0, y0=0, x1=1, y1=1, line=dict(color="rgba(0, 0, 0, 0.6)", width=2))
        ])

        # Update x-axes for each subplot
        fig3.update_xaxes(nticks=4, range=[0, 1], showgrid=True, tickvals=[0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1], row=1, col=1)  # For Vclay
        fig3.update_xaxes(nticks=4, range=[0, 0.4], showgrid=True, tickvals=[0, 0.1, 0.2, 0.3, 0.4], row=1, col=2)  # For Porosity
        fig3.update_xaxes(nticks=4, range=[1, 0], showgrid=True, tickvals=[0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1], row=1, col=3)  # For Saturation
        fig3.update_xaxes(range=[1, 1], showgrid=False, row=1, col=4)  # For Vertical Line Plot

        # Update y-axis with range based on df_p['MD'] and reverse the y-axis
        fig3.update_yaxes(nticks=20, range=[df_p['MD'].max(), df_p['MD'].min()], showgrid=True)

        # Show plot
        st.plotly_chart(fig3)        
  

def page2():
    session_state=get_session_state()
    st.success('VCLAY ANALYSIS')
    tab1, tab2 = st.tabs(["V_CLAY", "CUTOFF_PLOTS"])
    with tab1:
        session_state.dict1 = {}
        session_state.interval_ = []
        dict_vclay_NTG={}
        dict_vclay_PVH={}
        for Name,df in session_state.dict_Rawdataframe.items():
            interval = df.iloc[1,0]-df.iloc[0,0]
            interval= np.abs(interval)
            session_state.interval_.append(interval)
            df_vclay = df.sort_values(by='Vcl')
            df_vclay['PVH_Cum'] = df_vclay['Pi'].cumsum()
            df_vclay['PVH_Percent']=df_vclay['PVH_Cum']/(df_vclay['PVH_Cum'].max())
            Vclay_cutoff = np.linspace(0, 1, 21)
            Vclay_cutoff = pd.DataFrame(Vclay_cutoff, columns=['cutoff'])
            Vclay_cutoff['frequency'] = Vclay_cutoff['cutoff'].apply(lambda x: (df_vclay['Vcl'] <= x).sum())
            Vclay_cutoff['pay']=Vclay_cutoff['frequency']*interval
            Vclay_cutoff['NTG']=Vclay_cutoff['pay']/(Vclay_cutoff['pay'].max())
            Vclay_cutoff = pd.merge_asof(Vclay_cutoff, df_vclay[['Vcl','PVH_Percent']], left_on='cutoff', right_on='Vcl', direction='forward')
            Vclay_cutoff.drop('Vcl', axis=1, inplace=True)
            Vclay_cutoff.fillna(1,inplace=True)
            list1=[]
            list1.append(df_vclay)
            list1.append(Vclay_cutoff)
            session_state.dict1[Name]=list1
            dict_vclay_NTG[Name] = list(Vclay_cutoff['NTG'])
            dict_vclay_PVH[Name] = list(Vclay_cutoff['PVH_Percent'])
        # Retrieve the first list from the dictionary values
        first = next(iter(session_state.dict1.values()))
        # Retrieve the first column from the first list
        cutoff_ = list(first[1]['cutoff'])
        df_cutoff = pd.DataFrame({ 'cutoff': cutoff_ })
        NTG_summary = pd.DataFrame(dict_vclay_NTG)
        PVH_summary = pd.DataFrame(dict_vclay_PVH)
        session_state.NTG = pd.concat([df_cutoff, NTG_summary], axis=1)
        session_state.PVH = pd.concat([df_cutoff, PVH_summary], axis=1)
        selected_option = st.selectbox('Choose the well', list(session_state.dict1.keys()))
        col1, col2  = st.columns((1, 1))
        with col1:    
            df = session_state.dict1[selected_option][1]
            st.write('Data:', df)
        with col2:
            cutoff_entry_form('Vclay',0.6,session_state.dict1)
            index=8
            for i, key in enumerate(session_state.dict_wells.keys()):
                if index<len(session_state.dict_wells[key]):
                    session_state.dict_wells[key][index]=session_state['Vclay_Cutoff_df'].iloc[i]['Vclay_Cutoff']
                else:
                    session_state.dict_wells[key].append(session_state['Vclay_Cutoff_df'].iloc[i]['Vclay_Cutoff'])
            


            if st.button('Save Data to Excel'):
                excel_data = convert_dfs_to_excel(session_state.dict1)
                st.download_button(label = "Download File",data = excel_data,file_name = "VCLAY_Data.xlsx",mime= 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            

    with tab2:
        # Plot PVH Percent vs. Cutoff
        fig1 = go.Figure()
        for key, value in session_state.dict1.items():
            trace_PVH = go.Scatter(x=session_state.dict1[key][1]['cutoff'], y=session_state.dict1[key][1]['PVH_Percent'], mode='lines+markers', name=key, visible=True)
            fig1.add_trace(trace_PVH)

        fig1.update_xaxes(title_text="Cutoff", tickmode='array', tickvals=session_state.dict1[key][1]['cutoff'], showgrid=True)
        fig1.update_yaxes(title_text="PVH Percent")
        fig1.update_layout(width=1000,height=700,plot_bgcolor='white',paper_bgcolor='white',shapes=[dict(type="rect",xref="paper",yref="paper",x0=0,y0=0,x1=1,y1=1,line=dict(color="rgba(0, 0, 0, 0.6)", width=2))])    
        st.plotly_chart(fig1)

        # Plot NTG vs. Cutoff
        fig2 = go.Figure()
        for key, value in session_state.dict1.items():
            trace_NTG = go.Scatter(x=session_state.dict1[key][1]['cutoff'], y=session_state.dict1[key][1]['NTG'], mode='lines+markers', name=key, visible=True)
            fig2.add_trace(trace_NTG)

        fig2.update_xaxes(title_text="Cutoff", tickmode='array', tickvals=session_state.dict1[key][1]['cutoff'], showgrid=True)
        fig2.update_yaxes(title_text="NTG")
        fig2.update_layout(width=1000,height=700,plot_bgcolor='white',paper_bgcolor='white',shapes=[dict(type="rect",xref="paper",yref="paper",x0=0,y0=0,x1=1,y1=1,line=dict(color="rgba(0, 0, 0, 0.6)",  width=2))])

        st.plotly_chart(fig2)

    


def page3():
    session_state=get_session_state()
    st.success('POROSITY ANALYSIS')
    session_state.dict2 = {}
    dict_pige_NTG={}
    dict_pige_PVH={}
    Dataframe_Post_Vclay = {}
    for i,((key,value),(Key1, value1)) in enumerate(zip(session_state.dict1.items(),session_state.dict_wells.items())):
        well_pige = session_state.dict1[key][0][session_state.dict1[key][0]['Vcl']<=value1[8]]
        Post_Vclay = well_pige[['MD','Pi','Sw','Vcl']]
        Post_Vclay = Post_Vclay.sort_values(by = 'MD')
        Dataframe_Post_Vclay[key] = Post_Vclay
        well_pige=well_pige.sort_values(by='Pi')
        well_pige['PVH_Cum'] = well_pige['Vcl'].cumsum().max()-well_pige['Vcl'].cumsum()
        well_pige['PVH_Percent']=well_pige['PVH_Cum']/(well_pige['PVH_Cum'].max())
        Pi_cutoff = np.linspace(0, 0.32, 17)
        Pi_cutoff = pd.DataFrame(Pi_cutoff, columns=['cutoff'])
        Pi_cutoff['frequency'] = Pi_cutoff['cutoff'].apply(lambda x: (well_pige['Pi'] < x).sum())
        Pi_cutoff['pay']=(Pi_cutoff['frequency'].max()-Pi_cutoff['frequency'])*session_state.interval_[i]
        Pi_cutoff['NTG']=Pi_cutoff['pay']/(Pi_cutoff['pay'].max())
        Pi_cutoff = pd.merge_asof(Pi_cutoff, well_pige[['Pi','PVH_Percent']], left_on='cutoff', right_on='Pi', direction='forward')
        Pi_cutoff.drop('Pi', axis=1, inplace=True)
        Pi_cutoff.fillna(0,inplace=True)
        list2=[]
        list2.append(well_pige)
        list2.append(Pi_cutoff)
        session_state.dict2[key]=list2
        dict_pige_NTG[key] = list(Pi_cutoff['NTG'])
        dict_pige_PVH[key] = list(Pi_cutoff['PVH_Percent'])

    # Retrieve the first list from the dictionary values
    first = next(iter(session_state.dict2.values()))
    # Retrieve the first column from the first list
    cutoff_ = list(first[1]['cutoff'])
    df_cutoff = pd.DataFrame({ 'cutoff': cutoff_ })
    NTG_summary = pd.DataFrame(dict_pige_NTG)
    PVH_summary = pd.DataFrame(dict_pige_PVH)
    session_state.NTG = pd.concat([df_cutoff, NTG_summary], axis=1)
    session_state.PVH = pd.concat([df_cutoff, PVH_summary], axis=1)
    tab1, tab2, tab3 = st.tabs(["POST_VCLAY_DATA","POROSITY", "CUTOFF_PLOTS"])
    with tab1:
        Histogram(Dataframe_Post_Vclay)
    with tab2:
        selected_option = st.selectbox('Choose the well for Porosity', list(session_state.dict2.keys()))
        col1, col2  = st.columns((1, 1))
        with col1:    
            df = session_state.dict2[selected_option][1]
            st.write('Data:', df)
        with col2:
            cutoff_entry_form('Porosity',0.06,session_state.dict2)
            index=9
            for i, key in enumerate(session_state.dict_wells.keys()):
                if index<len(session_state.dict_wells[key]):
                    session_state.dict_wells[key][index]=session_state['Porosity_Cutoff_df'].iloc[i]['Porosity_Cutoff']
                else:
                    session_state.dict_wells[key].append(session_state['Porosity_Cutoff_df'].iloc[i]['Porosity_Cutoff'])
                


            if st.button('Save Data to Excel'):
                excel_data = convert_dfs_to_excel(session_state.dict2)
                st.download_button(label = "Download File",data = excel_data,file_name = "Porosity_Data.xlsx",mime= 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    with tab3:
        # Plot PVH Percent vs. Cutoff
        fig1 = go.Figure()
        for key, value in session_state.dict2.items():
            trace_PVH = go.Scatter(x=session_state.dict2[key][1]['cutoff'], y=session_state.dict2[key][1]['PVH_Percent'], mode='lines+markers', name=key, visible=True)
            fig1.add_trace(trace_PVH)

        fig1.update_xaxes(title_text="Cutoff", tickmode='array', tickvals=session_state.dict2[key][1]['cutoff'], showgrid=True)
        fig1.update_yaxes(title_text="PVH Percent")
        fig1.update_layout(width=1000,height=700,plot_bgcolor='white',paper_bgcolor='white',shapes=[dict(type="rect",xref="paper",yref="paper",x0=0,y0=0,x1=1,y1=1,line=dict(color="rgba(0, 0, 0, 0.6)", width=2))])    
        st.plotly_chart(fig1)

        # Plot NTG vs. Cutoff
        fig2 = go.Figure()
        for key, value in session_state.dict2.items():
            trace_NTG = go.Scatter(x=session_state.dict2[key][1]['cutoff'], y=session_state.dict2[key][1]['NTG'], mode='lines+markers', name=key, visible=True)
            fig2.add_trace(trace_NTG)

        fig2.update_xaxes(title_text="Cutoff", tickmode='array', tickvals=session_state.dict2[key][1]['cutoff'], showgrid=True)
        fig2.update_yaxes(title_text="NTG")
        fig2.update_layout(width=1000,height=700,plot_bgcolor='white',paper_bgcolor='white',shapes=[dict(type="rect",xref="paper",yref="paper",x0=0,y0=0,x1=1,y1=1,line=dict(color="rgba(0, 0, 0, 0.6)",  width=2))])

        st.plotly_chart(fig2)


def page4():
    session_state=get_session_state()
    st.success('SATURATION ANALYSIS')
    session_state.dict3 = {}
    dict_sw_NTG={}
    dict_sw_PVH={}
    Dataframe_Post_Porosity = {}

    for i,((key,value),(Key1, value1)) in enumerate(zip(session_state.dict2.items(),session_state.dict_wells.items())):
        well_sw = session_state.dict2[key][0][session_state.dict2[key][0]['Pi']>=value1[9]]
        Post_Porosity = well_sw[['MD','Pi','Sw','Vcl']]
        Post_Porosity = Post_Porosity.sort_values(by = 'MD')
        Dataframe_Post_Porosity[key] = Post_Porosity
        well_sw=well_sw.sort_values(by='Sw')
        well_sw['HCPV']=well_sw['Pi']*(1-well_sw['Sw'])
        well_sw['PVH_Cum'] = well_sw['HCPV'].cumsum()
        well_sw['PVH_Percent']=well_sw['PVH_Cum']/(well_sw['PVH_Cum'].max())
        sw_cutoff = np.linspace(0, 1, 21)
        sw_cutoff = pd.DataFrame(sw_cutoff, columns=['cutoff'])
        sw_cutoff['frequency'] = sw_cutoff['cutoff'].apply(lambda x: (well_sw['Sw'] <= x).sum())
        sw_cutoff['pay']=sw_cutoff['frequency']*session_state.interval_[i]
        sw_cutoff['NTG']=sw_cutoff['pay']/(sw_cutoff['pay'].max())
        sw_cutoff = pd.merge_asof(sw_cutoff, well_sw[['Sw','PVH_Percent']], left_on='cutoff', right_on='Sw', direction='forward')
        sw_cutoff.drop('Sw', axis=1, inplace=True)
        sw_cutoff.fillna(1,inplace=True)
        if (well_sw['Sw'] == 1).all():
            sw_cutoff['PVH_Percent'] = 0
        list3=[]
        list3.append(well_sw)
        list3.append(sw_cutoff)
        session_state.dict3[key]=list3
        dict_sw_NTG[key] = list(sw_cutoff['NTG'])
        dict_sw_PVH[key] = list(sw_cutoff['PVH_Percent'])
    # Retrieve the first list from the dictionary values
    first = next(iter(session_state.dict3.values()))
    # Retrieve the first column from the first list
    cutoff_ = list(first[1]['cutoff'])
    df_cutoff = pd.DataFrame({ 'cutoff': cutoff_ })
    NTG_summary = pd.DataFrame(dict_sw_NTG)
    PVH_summary = pd.DataFrame(dict_sw_PVH)
    session_state.NTG = pd.concat([df_cutoff, NTG_summary], axis=1)
    session_state.PVH = pd.concat([df_cutoff, PVH_summary], axis=1)
    tab1, tab2, tab3 = st.tabs(["POST_POROSITY_DATA","SATURATION", "CUTOFF_PLOTS"])
    with tab1:
        Histogram(Dataframe_Post_Porosity)
    with tab2:
        selected_option = st.selectbox('Choose the well for Saturation', list(session_state.dict3.keys()))
        col1, col2  = st.columns((1, 1))
        with col1:    
            df = session_state.dict3[selected_option][1]
            st.write('Data:', df)
        with col2:
            cutoff_entry_form('Saturation',0.8,session_state.dict3)
            index=10
            for i, key in enumerate(session_state.dict_wells.keys()):
                if index<len(session_state.dict_wells[key]):
                    session_state.dict_wells[key][index]=session_state['Saturation_Cutoff_df'].iloc[i]['Saturation_Cutoff']
                else:
                    session_state.dict_wells[key].append(session_state['Saturation_Cutoff_df'].iloc[i]['Saturation_Cutoff'])
                


            if st.button('Save Data to Excel'):
                excel_data = convert_dfs_to_excel(session_state.dict3)
                st.download_button(label = "Download File",data = excel_data,file_name = "Saturation_Data.xlsx",mime= 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    with tab3:
        # Plot PVH Percent vs. Cutoff
        fig1 = go.Figure()
        for key, value in session_state.dict3.items():
            trace_PVH = go.Scatter(x=session_state.dict3[key][1]['cutoff'], y=session_state.dict3[key][1]['PVH_Percent'], mode='lines+markers', name=key, visible=True)
            fig1.add_trace(trace_PVH)

        fig1.update_xaxes(title_text="Cutoff", tickmode='array', tickvals=session_state.dict3[key][1]['cutoff'], showgrid=True)
        fig1.update_yaxes(title_text="PVH Percent")
        fig1.update_layout(width=1000,height=700,plot_bgcolor='white',paper_bgcolor='white',shapes=[dict(type="rect",xref="paper",yref="paper",x0=0,y0=0,x1=1,y1=1,line=dict(color="rgba(0, 0, 0, 0.6)", width=2))])    
        st.plotly_chart(fig1)

        # Plot NTG vs. Cutoff
        fig2 = go.Figure()
        for key, value in session_state.dict3.items():
            trace_NTG = go.Scatter(x=session_state.dict3[key][1]['cutoff'], y=session_state.dict3[key][1]['NTG'], mode='lines+markers', name=key, visible=True)
            fig2.add_trace(trace_NTG)

        fig2.update_xaxes(title_text="Cutoff", tickmode='array', tickvals=session_state.dict3[key][1]['cutoff'], showgrid=True)
        fig2.update_yaxes(title_text="NTG")
        fig2.update_layout(width=1000,height=700,plot_bgcolor='white',paper_bgcolor='white',shapes=[dict(type="rect",xref="paper",yref="paper",x0=0,y0=0,x1=1,y1=1,line=dict(color="rgba(0, 0, 0, 0.6)",  width=2))])

        st.plotly_chart(fig2)

def page5():
    
    st.success('PAY AND PARAMETERS SUMMARY')
    session_state=get_session_state()    
    
    tab1, tab2, tab3 = st.tabs(["POST_CUTOFF_DATA", "PAY AND PARAMETERS SUMMARY","QC_Plots"])
   
    def main_calculation():
        session_state=get_session_state()
        session_state.dict4 = {}
        session_state.dict5 = {}
        session_state.dict_Pay_Parameter={} 
        def pay_calculator(df):
            if Dev_dataframe.empty:
                dummy_OWC= df
                dummy_OWC['DEVIATION_ANGLE']= 0
                dummy_OWC['TVD'] = session_state.interval_[i]       
            else:
                dummy_OWC = pd.merge_asof(df, Dev_dataframe[['MD','DEVIATION_ANGLE']], on='MD', direction='nearest')
                dummy_OWC['TVD'] = session_state.interval_[i]* np.cos(np.deg2rad(dummy_OWC['DEVIATION_ANGLE']))
            pay = dummy_OWC['TVD'].sum()
            return pay
        
        for i,((key,value),(Key1, value1)) in enumerate(zip(session_state.dict3.items(),session_state.dict_wells.items())):
            Top = value1[0]
            GOC=  value1[1]
            OWC= value1[2]
            Bottom=value1[3]
            S_W_Cut = value1[10]
            R_Type = value1[6]
            Dev_dataframe = value1[7]
            summary = []   
            if R_Type=='O':     
                well_postcutoff_OWC = session_state.dict3[key][0][(session_state.dict3[key][0]['MD'] >= Top) & (session_state.dict3[key][0]['MD'] <= OWC)]
                Gross_pay_Oil_df=well_postcutoff_OWC.sort_values(by = 'MD')
                Gross_Pay_Oil = pay_calculator(Gross_pay_Oil_df)
                well_postcutoff_OWC = well_postcutoff_OWC[well_postcutoff_OWC['Sw']<=S_W_Cut]
                well_postcutoff_OWC =well_postcutoff_OWC.sort_values(by='MD')
                post_cutoff_df = well_postcutoff_OWC
                OWC = well_postcutoff_OWC['MD'].iloc[-1]
                Net_Pay_Oil = pay_calculator(well_postcutoff_OWC)
                Porosity_Oil = well_postcutoff_OWC['Pi'].mean()
                Saturation_Oil = well_postcutoff_OWC['Sw'].mean()
                well_postcutoff_bottom = session_state.dict3[key][0][(session_state.dict3[key][0]['MD'] > OWC) & (session_state.dict3[key][0]['MD'] <= Bottom)]
                well_postcutoff_bottom =well_postcutoff_bottom.sort_values(by='MD')
                blank_row = pd.DataFrame({}, index=[0])
                well_postcutoff = pd.concat([well_postcutoff_OWC,blank_row, well_postcutoff_bottom]).reset_index(drop=True)
                summary_list = [Gross_Pay_Oil,Net_Pay_Oil,Porosity_Oil,Saturation_Oil,0,0,0,0]
                summary.extend(summary_list)
                session_state.dict_Pay_Parameter[key]=summary

                
            
            elif R_Type=='G':
                well_postcutoff_GOC = session_state.dict3[key][0][(session_state.dict3[key][0]['MD'] >= Top) & (session_state.dict3[key][0]['MD'] <= GOC)]
                Gross_pay_gas_df=well_postcutoff_GOC.sort_values(by = 'MD')
                Gross_Pay_gas = pay_calculator(Gross_pay_gas_df)
                well_postcutoff_GOC = well_postcutoff_GOC[well_postcutoff_GOC['Sw']<=S_W_Cut]
                well_postcutoff_GOC =well_postcutoff_GOC.sort_values(by='MD')
                post_cutoff_df = well_postcutoff_GOC
                GOC = well_postcutoff_GOC['MD'].iloc[-1]
                Net_Pay_gas = pay_calculator(well_postcutoff_GOC)
                Porosity_gas = well_postcutoff_GOC['Pi'].mean()
                Saturation_gas = well_postcutoff_GOC['Sw'].mean()
                well_postcutoff_bottom = session_state.dict3[key][0][(session_state.dict3[key][0]['MD'] > GOC) & (session_state.dict3[key][0]['MD'] <= Bottom)]
                well_postcutoff_bottom =well_postcutoff_bottom.sort_values(by='MD')
                blank_row = pd.DataFrame({}, index=[0])
                well_postcutoff = pd.concat([well_postcutoff_GOC,blank_row, well_postcutoff_bottom]).reset_index(drop=True)
                summary_list = [0,0,0,0,Gross_Pay_gas,Net_Pay_gas,Porosity_gas,Saturation_gas]
                summary.extend(summary_list)
                session_state.dict_Pay_Parameter[key]=summary
            
                
            elif R_Type=='OG':
                if GOC==0:                
                    well_postcutoff_OWC = session_state.dict3[key][0][(session_state.dict3[key][0]['MD'] >= Top) & (session_state.dict3[key][0]['MD'] <= OWC)]
                    Gross_pay_Oil_df=well_postcutoff_OWC.sort_values(by = 'MD')
                    Gross_Pay_Oil = pay_calculator(Gross_pay_Oil_df)
                    well_postcutoff_OWC = well_postcutoff_OWC[well_postcutoff_OWC['Sw']<=S_W_Cut]
                    well_postcutoff_OWC =well_postcutoff_OWC.sort_values(by='MD')
                    post_cutoff_df = well_postcutoff_OWC
                    OWC = well_postcutoff_OWC['MD'].iloc[-1]
                    Net_Pay_Oil = pay_calculator(well_postcutoff_OWC)
                    Porosity_Oil = well_postcutoff_OWC['Pi'].mean()
                    Saturation_Oil = well_postcutoff_OWC['Sw'].mean()
                    well_postcutoff_bottom = session_state.dict3[key][0][(session_state.dict3[key][0]['MD'] > OWC) & (session_state.dict3[key][0]['MD'] <= Bottom)]
                    well_postcutoff_bottom =well_postcutoff_bottom.sort_values(by='MD')                
                    blank_row = pd.DataFrame({}, index=[0])
                    well_postcutoff = pd.concat([well_postcutoff_OWC,blank_row, well_postcutoff_bottom]).reset_index(drop=True)
                    summary_list = [Gross_Pay_Oil,Net_Pay_Oil,Porosity_Oil,Saturation_Oil,0,0,0,0]
                    summary.extend(summary_list)
                    session_state.dict_Pay_Parameter[key]=summary
            
                    
                elif OWC==0:
                    well_postcutoff_GOC = session_state.dict3[key][0][(session_state.dict3[key][0]['MD'] >= Top) & (session_state.dict3[key][0]['MD'] <= GOC)]
                    Gross_pay_gas_df=well_postcutoff_GOC.sort_values(by = 'MD')
                    Gross_Pay_gas = pay_calculator(Gross_pay_gas_df)
                    well_postcutoff_GOC = well_postcutoff_GOC[well_postcutoff_GOC['Sw']<=S_W_Cut]
                    well_postcutoff_GOC =well_postcutoff_GOC.sort_values(by='MD')
                    post_cutoff_df = well_postcutoff_GOC
                    GOC = well_postcutoff_GOC['MD'].iloc[-1]
                    Net_Pay_gas = pay_calculator(well_postcutoff_GOC)
                    Porosity_gas = well_postcutoff_GOC['Pi'].mean()
                    Saturation_gas = well_postcutoff_GOC['Sw'].mean()
                    well_postcutoff_bottom = session_state.dict3[key][0][(session_state.dict3[key][0]['MD'] > GOC) & (session_state.dict3[key][0]['MD'] <= Bottom)]
                    well_postcutoff_bottom =well_postcutoff_bottom.sort_values(by='MD')
                    blank_row = pd.DataFrame({}, index=[0])
                    well_postcutoff = pd.concat([well_postcutoff_GOC,blank_row, well_postcutoff_bottom]).reset_index(drop=True)
                    summary_list = [0,0,0,0,Gross_Pay_gas,Net_Pay_gas,Porosity_gas,Saturation_gas]
                    summary.extend(summary_list)
                    session_state.dict_Pay_Parameter[key]=summary
                            
                else:
                    
                    well_postcutoff_GOC = session_state.dict3[key][0][(session_state.dict3[key][0]['MD'] >= Top) & (session_state.dict3[key][0]['MD'] <= GOC)]
                    Gross_pay_gas_df=well_postcutoff_GOC.sort_values(by = 'MD')
                    Gross_Pay_gas = pay_calculator(Gross_pay_gas_df)
                    well_postcutoff_GOC = well_postcutoff_GOC[well_postcutoff_GOC['Sw']<=S_W_Cut]
                    well_postcutoff_GOC =well_postcutoff_GOC.sort_values(by='MD')
                    post_cutoff_df_1 = well_postcutoff_GOC
                    Net_Pay_gas = pay_calculator(well_postcutoff_GOC)
                    Porosity_gas = well_postcutoff_GOC['Pi'].mean()
                    Saturation_gas = well_postcutoff_GOC['Sw'].mean()
                    blank_row = pd.DataFrame({}, index=[0])
                    well_postcutoff_OWC = session_state.dict3[key][0][(session_state.dict3[key][0]['MD'] > GOC) & (session_state.dict3[key][0]['MD'] <= OWC)]
                    Gross_pay_Oil_df=well_postcutoff_OWC.sort_values(by = 'MD')
                    Gross_Pay_Oil = pay_calculator(Gross_pay_Oil_df)
                    well_postcutoff_OWC = well_postcutoff_OWC[well_postcutoff_OWC['Sw']<=S_W_Cut]
                    well_postcutoff_OWC =well_postcutoff_OWC.sort_values(by='MD')
                    post_cutoff_df_2 = well_postcutoff_OWC
                    post_cutoff_df = pd.concat([post_cutoff_df_1,post_cutoff_df_2]).reset_index(drop=True)
                    OWC = well_postcutoff_OWC['MD'].iloc[-1]
                    Net_Pay_Oil = pay_calculator(well_postcutoff_OWC)
                    Porosity_Oil = well_postcutoff_OWC['Pi'].mean()
                    Saturation_Oil = well_postcutoff_OWC['Sw'].mean()
                    blank_row_1 = pd.DataFrame({}, index=[0])
                    well_postcutoff_bottom = session_state.dict3[key][0][(session_state.dict3[key][0]['MD'] > OWC) & (session_state.dict3[key][0]['MD'] <= Bottom)]
                    well_postcutoff_bottom =well_postcutoff_bottom.sort_values(by='MD')
                    well_postcutoff = pd.concat([well_postcutoff_GOC,blank_row,well_postcutoff_OWC,blank_row_1, well_postcutoff_bottom]).reset_index(drop=True)
                    summary_list = [Gross_Pay_Oil ,Net_Pay_Oil,Porosity_Oil,Saturation_Oil,Gross_Pay_gas,Net_Pay_gas,Porosity_gas,Saturation_gas]
                    summary.extend(summary_list)
                    session_state.dict_Pay_Parameter[key]=summary

            session_state.dict4[key]=well_postcutoff.iloc[:,0:4]
            session_state.dict5[key]=post_cutoff_df.iloc[:,0:4]
        session_state.df_final=pd.DataFrame.from_dict(session_state.dict_Pay_Parameter,orient='index')
        session_state.df_final.columns= ['Gross_Pay_Oil' ,'Net_Pay_Oil','Porosity_Oil','Saturation_Oil','Gross_Pay_Gas','Net_Pay_Gas','Porosity_Gas','Saturation_Gas']
        return session_state.df_final
    main_calculation()
        
    with tab1:
        
        selected_option = st.selectbox('Choose the well', list(session_state.dict4.keys()))
        col1, col2  = st.columns((2, 2))
        with col1:    
            df = session_state.dict4[selected_option]
            st.write('Data:', df)
        with col2:
            dee = session_state.dict5[selected_option]
            de = dee.describe()
            st.write('Data_Statistics:',de)
        if st.button('Display Histograms'):    
            columns_to_plot = dee.columns[1:4]
            for column in columns_to_plot:
                dee[column] = pd.to_numeric(dee[column], errors='coerce')
            dee = dee.replace([np.inf, -np.inf], np.nan).dropna(subset=columns_to_plot)
            fig, axs = plt.subplots(1, 3, figsize=(15, 5))  
            for ax, column in zip(axs, columns_to_plot):
                ax.hist(dee[column], bins=10, edgecolor='black')
                ax.set_title(f'Histogram of {column}')
                ax.set_xlabel(column)
                ax.set_ylabel('Frequency')
            st.pyplot(fig)
    with tab2:           
        st.write(session_state.df_final)   

        def load_workbook_from_file(file):
            """Loads an Excel workbook from a file-like object."""
            return load_workbook(io.BytesIO(file.read()), data_only=False)

        def save_workbook_to_bytes(wb):
            """Saves a workbook to a BytesIO object."""
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output

        def force_recalculate(wb):
            """Force Excel to recalculate formulas when the workbook is opened."""
            wb.properties.calcMode = 'auto'
            wb.properties.calcId = 0

        def replace_in_row(sheet, row_number, old_text, new_text):
            """Replace occurrences of old_text with new_text in a specific row."""
            for cell in sheet[row_number]:
                if cell.value and isinstance(cell.value, str) and old_text in cell.value:
                    cell.value = cell.value.replace(old_text, new_text)

        uploaded_file = st.file_uploader("Choose the Post_cutoff excel file", type='xlsx')
        if uploaded_file is not None:
            try:
                wb = load_workbook_from_file(uploaded_file)

                # Log existing sheet names for debugging
                existing_sheets = wb.sheetnames

                # Assuming session_state.dict4 contains the new sheet names
                sheet_name_mapping = {}
                for i, ((key, value), (Key1, value1)) in enumerate(zip(session_state.dict4.items(), session_state.dict_wells.items())):
                    old_name = f'Sheet{i+1}'
                    new_name = key
                    sheet_name_mapping[old_name] = new_name

                    if old_name in wb.sheetnames:
                        sheet = wb[old_name]
                        sheet.title = new_name
                    else:
                        st.error(f"Sheet '{old_name}' does not exist in the workbook.")

                    # Clear range A2:D*
                    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=4):
                        for cell in row:
                            cell.value = None

                    # Set new values in range A2:D*
                    for row_idx, row_value in enumerate(session_state.dict4[key].values.tolist(), start=2):
                        for col_idx, cell_value in enumerate(row_value, start=1):
                            sheet.cell(row=row_idx, column=col_idx, value=cell_value)

                    sheet['P2'].value = round(session_state.interval_[i], 4)
                    sheet['O2'].value = key
                    sheet['O3'].value = 'Reservoir_Type'
                    sheet['P3'].value = value1[6]
                    sheet['O4'].value = 'GOC'
                    sheet['P4'].value = value1[1]
                    sheet['O5'].value = 'OWC'
                    sheet['P5'].value = value1[2]
                    sheet['S21'].value = value1[4]
                    sheet['T21'].value = value1[5]
                    sheet['P6'].value = session_state.dict_Pay_Parameter[key][4]
                    sheet['P7'].value = session_state.dict_Pay_Parameter[key][0]
                    sheet['P8'].value = value1[0]
                    sheet['P9'].value = value1[3]

                    # Set new values in range I24:K*
                    for row_idx, row_value in enumerate(value1[7].values.tolist(), start=24):
                        for col_idx, cell_value in enumerate(row_value, start=9):
                            sheet.cell(row=row_idx, column=col_idx, value=cell_value)
                # Update references in the summary sheet
                summary_sheet = wb['Summary']
                for i,(old_name, new_name) in enumerate(sheet_name_mapping.items()):
                    replace_in_row(summary_sheet, i+4, old_name, new_name)
                # Force Excel to recalculate formulas
                force_recalculate(wb)

                updated_file = save_workbook_to_bytes(wb)
                st.download_button(
                    label="Download File",
                    data=updated_file,
                    file_name="postcutoff_Data.xlsx",
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            except Exception as e:
                st.error(f"An error occurred: {e}")
    
    with tab3:
        def y_values_creator(data_dict):
            list = []
            for key,value in data_dict.items():
                if data_dict is session_state["dict5"]:
                    well_pige=value
                else:
                    well_pige=value[0]
                Post_Vclay = well_pige[['MD']]
                Post_Vclay = Post_Vclay.sort_values(by = 'MD')
                list.append(Post_Vclay['MD'].values)
            return list
        def segements_creator(data_dict,x,color_,interval_list):
            color_segment=color_
            list_segment=y_values_creator(data_dict)
            dict_segment = {}
            for i,(key,value) in enumerate(data_dict.items()):
                x_value = x 
                segments = []
                interval_seg=interval_list[i]
                y_values = list_segment[i]
                for i in range(len(y_values) - 1):
                    if abs(y_values[i + 1] - y_values[i] - interval_seg) < 1e-10:  
                        segments.append(go.Scatter(x=[x_value, x_value], y=[y_values[i], y_values[i + 1]],mode='lines', line=dict(color=color_segment, width=15),showlegend=False))
                dict_segment[key]=segments
            return dict_segment
    
        segment_vclay=segements_creator(session_state.dict2,1,'blue',session_state.interval_)
        segment_porosity=segements_creator(session_state.dict3,2,'green',session_state.interval_)
        segment_saturation=segements_creator(session_state.dict5,3,'red',session_state.interval_)
        
        df_plotted = st.selectbox('Choos the well for plot', list(session_state.dict_Rawdataframe.keys()))
        df_p = session_state.dict_Rawdataframe[df_plotted]
        # Create the figure with subplots
        fig3 = make_subplots(rows=1, cols=4, shared_yaxes=True, column_widths=[0.25, 0.25, 0.25, 0.25])

        # Add scatter plots for each column
        fig3.add_trace(go.Scatter(x=df_p['Vcl'], y=df_p['MD'], mode='lines', name='Vclay'), row=1, col=1)
        fig3.add_trace(go.Scatter(x=df_p['Pi'], y=df_p['MD'], mode='lines', name='Porosity',fill='tozerox',fillcolor='rgba(0, 255, 0, 0.1)'), row=1, col=2)
        fig3.add_trace(go.Scatter(x=1-df_p['Sw'], y=df_p['MD'], mode='lines', name='Saturation', fill='tozerox',fillcolor='rgba(255, 0, 0, 0.1)'), row=1, col=3)

        # Add the vertical line segments as the fourth subplot
        for segment in segment_vclay[df_plotted]:
            fig3.add_trace(segment,  row=1, col=4)
        for segment in segment_porosity[df_plotted]:
            fig3.add_trace(segment,  row=1, col=4)
        for segment in segment_saturation[df_plotted]:
            fig3.add_trace(segment,  row=1, col=4)
        
        y_value_OWC = session_state.dict_wells[df_plotted][2]
        y_value_GOC = session_state.dict_wells[df_plotted][1]   
        fig3.add_hline(y=y_value_OWC,line=dict(color="red", width=2),row="all", col="all")
        fig3.add_hline(y=y_value_GOC,line=dict(color="green", width=2),row="all", col="all")

        # Update layout
        fig3.update_layout(title_text="CUTOFF_PROPERTIES", height=800, width=1200, shapes=[
            dict(type="rect", xref="paper", yref="paper", x0=0, y0=0, x1=1, y1=1, line=dict(color="rgba(0, 0, 0, 0.6)", width=2))
        ])

        # Update x-axes for each subplot
        fig3.update_xaxes(nticks=4, range=[0, 1], showgrid=True, tickvals=[0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1], row=1, col=1)  # For Vclay
        fig3.update_xaxes(nticks=4, range=[0, 0.4], showgrid=True, tickvals=[0, 0.1, 0.2, 0.3, 0.4], row=1, col=2)  # For Porosity
        fig3.update_xaxes(nticks=4, range=[0,1], showgrid=True, tickvals=[0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1], row=1, col=3)  # For Saturation
        fig3.update_xaxes(range=[0, 4], showgrid=False, row=1, col=4)  # For Vertical Line Plot

        # Update y-axis with range based on df_p['MD'] and reverse the y-axis
        fig3.update_yaxes(nticks=20, range=[df_p['MD'].max(), df_p['MD'].min()], showgrid=True)

        # Show plot
        st.plotly_chart(fig3)

def page6():
    st.success('SENSITIVITY ANALYSIS')
    session_state=get_session_state()    
    
    tab1, tab2 = st.tabs(["SENSITIVITY_DATA", "SENSITIVITY_PLOT"])
    with tab1:
        df = pd.DataFrame('', index=range(20), columns=["VCLAY_CUTOFF" ,"POROSITY_CUTOFF" , "SATURATION_CUTOFF"])
        # Create a GridOptionsBuilder object
        gb = GridOptionsBuilder.from_dataframe(df)
        gb.configure_default_column(editable=True)
        gb.configure_grid_options(enableRangeSelection=True)
        grid_options = gb.build()
        grid_response = AgGrid(df, gridOptions=grid_options, editable=True)

            # Get the updated DataFrame after user edits
        session_state.updated_df = grid_response['data']
           
        if st.button("Submit"):
            session_state.updated_df = session_state.updated_df.dropna(how='all').replace('', pd.NA).dropna(how='all')
            session_state.updated_df=session_state.updated_df.drop_duplicates()
            session_state.No_of_rows=session_state.updated_df.shape[0]
            st.write(session_state.updated_df)
            


            def pay_calculator(df):
                        if Dev_dataframe.empty:
                            dummy_OWC= df
                            dummy_OWC['DEVIATION_ANGLE']= 0
                            dummy_OWC['TVD'] = session_state.interval_[i]       
                        else:
                            dummy_OWC = pd.merge_asof(df, Dev_dataframe[['MD','DEVIATION_ANGLE']], on='MD', direction='nearest')
                            dummy_OWC['TVD'] = session_state.interval_[i]* np.cos(np.deg2rad(dummy_OWC['DEVIATION_ANGLE']))
                        pay = dummy_OWC['TVD'].sum()
                        return pay

            session_state.cases = {}
            session_state.Average_Property = {}
            for row in session_state.updated_df.itertuples(index=True, name='Pandas'):
                inde = int(row.Index)+1
                Vclay_Cutoff = float(row.VCLAY_CUTOFF)
                Porosity_Cutoff = float(row.POROSITY_CUTOFF)
                Saturation_Cutoff = float(row.SATURATION_CUTOFF)
                session_state.dict_Pay_P = {}
                summary_Average_property = []
                for i,((key,value),(Key1, value1)) in enumerate(zip(session_state.dict_Rawdataframe.items(),session_state.dict_wells.items())):
                        Top = value1[0]
                        GOC=  value1[1]
                        OWC= value1[2]
                        Bottom=value1[3]
                        R_Type = value1[6]
                        Dev_dataframe = value1[7]
                        summary = []   
                        if R_Type=='O':     
                            well_postcutoff_OWC = session_state.dict_Rawdataframe[key][(session_state.dict_Rawdataframe[key]['MD'] >= Top) & (session_state.dict_Rawdataframe[key]['MD'] <= OWC)]
                            well_postcutoff_OWC = well_postcutoff_OWC[(well_postcutoff_OWC['Vcl'] <= Vclay_Cutoff) & (well_postcutoff_OWC['Pi'] >= Porosity_Cutoff)] 
                            Gross_pay_Oil_df=well_postcutoff_OWC.sort_values(by = 'MD')
                            Gross_Pay_Oil = pay_calculator(Gross_pay_Oil_df)
                            well_postcutoff_OWC = well_postcutoff_OWC[well_postcutoff_OWC['Sw']<=Saturation_Cutoff]
                            well_postcutoff_OWC =well_postcutoff_OWC.sort_values(by='MD')
                            Net_Pay_Oil = pay_calculator(well_postcutoff_OWC)
                            Porosity_Oil = well_postcutoff_OWC['Pi'].mean()
                            Saturation_Oil = well_postcutoff_OWC['Sw'].mean()
                            summary_list = [Gross_Pay_Oil,Net_Pay_Oil,Porosity_Oil,Saturation_Oil,0,0,0,0]
                            summary.extend(summary_list)
                            session_state.dict_Pay_P[key]=summary

                        elif R_Type=='G':
                            well_postcutoff_GOC = session_state.dict_Rawdataframe[key][(session_state.dict_Rawdataframe[key]['MD'] >= Top) & (session_state.dict_Rawdataframe[key]['MD'] <= GOC)]
                            well_postcutoff_GOC = well_postcutoff_GOC[(well_postcutoff_GOC['Vcl'] <= Vclay_Cutoff) & (well_postcutoff_GOC['Pi'] >= Porosity_Cutoff)]
                            Gross_pay_gas_df=well_postcutoff_GOC.sort_values(by = 'MD')
                            Gross_Pay_gas = pay_calculator(Gross_pay_gas_df)
                            well_postcutoff_GOC = well_postcutoff_GOC[well_postcutoff_GOC['Sw']<=Saturation_Cutoff]
                            well_postcutoff_GOC =well_postcutoff_GOC.sort_values(by='MD')
                            Net_Pay_gas = pay_calculator(well_postcutoff_GOC)
                            Porosity_gas = well_postcutoff_GOC['Pi'].mean()
                            Saturation_gas = well_postcutoff_GOC['Sw'].mean()
                            summary_list = [0,0,0,0,Gross_Pay_gas,Net_Pay_gas,Porosity_gas,Saturation_gas]
                            summary.extend(summary_list)
                            session_state.dict_Pay_P[key]=summary
                        
                            
                        elif R_Type=='OG':
                            if GOC==0:                
                                well_postcutoff_OWC = session_state.dict_Rawdataframe[key][(session_state.dict_Rawdataframe[key]['MD'] >= Top) & (session_state.dict_Rawdataframe[key]['MD'] <= OWC)]
                                well_postcutoff_OWC = well_postcutoff_OWC[(well_postcutoff_OWC['Vcl'] <= Vclay_Cutoff) & (well_postcutoff_OWC['Pi'] >= Porosity_Cutoff)]
                                Gross_pay_Oil_df=well_postcutoff_OWC.sort_values(by = 'MD')
                                Gross_Pay_Oil = pay_calculator(Gross_pay_Oil_df)
                                well_postcutoff_OWC = well_postcutoff_OWC[well_postcutoff_OWC['Sw']<=Saturation_Cutoff]
                                well_postcutoff_OWC =well_postcutoff_OWC.sort_values(by='MD')
                                Net_Pay_Oil = pay_calculator(well_postcutoff_OWC)
                                Porosity_Oil = well_postcutoff_OWC['Pi'].mean()
                                Saturation_Oil = well_postcutoff_OWC['Sw'].mean()
                                summary_list = [Gross_Pay_Oil,Net_Pay_Oil,Porosity_Oil,Saturation_Oil,0,0,0,0]
                                summary.extend(summary_list)
                                session_state.dict_Pay_P[key]=summary
                        
                                
                            elif OWC==0:
                                well_postcutoff_GOC = session_state.dict_Rawdataframe[key][(session_state.dict_Rawdataframe[key]['MD'] >= Top) & (session_state.dict_Rawdataframe[key]['MD'] <= GOC)]
                                well_postcutoff_GOC = well_postcutoff_GOC[(well_postcutoff_GOC['Vcl'] <= Vclay_Cutoff) & (well_postcutoff_GOC['Pi'] >= Porosity_Cutoff)]
                                Gross_pay_gas_df=well_postcutoff_GOC.sort_values(by = 'MD')
                                Gross_Pay_gas = pay_calculator(Gross_pay_gas_df)
                                well_postcutoff_GOC = well_postcutoff_GOC[well_postcutoff_GOC['Sw']<=Saturation_Cutoff]
                                well_postcutoff_GOC =well_postcutoff_GOC.sort_values(by='MD')
                                Net_Pay_gas = pay_calculator(well_postcutoff_GOC)
                                Porosity_gas = well_postcutoff_GOC['Pi'].mean()
                                Saturation_gas = well_postcutoff_GOC['Sw'].mean()
                                summary_list = [0,0,0,0,Gross_Pay_gas,Net_Pay_gas,Porosity_gas,Saturation_gas]
                                summary.extend(summary_list)
                                session_state.dict_Pay_P[key]=summary
                                            
                            else:
                                well_postcutoff_GOC = session_state.dict_Rawdataframe[key][(session_state.dict_Rawdataframe[key]['MD'] >= Top) & (session_state.dict_Rawdataframe[key]['MD'] <= GOC)]
                                well_postcutoff_GOC = well_postcutoff_GOC[(well_postcutoff_GOC['Vcl'] <= Vclay_Cutoff) & (well_postcutoff_GOC['Pi'] >= Porosity_Cutoff)]
                                Gross_pay_gas_df=well_postcutoff_GOC.sort_values(by = 'MD')
                                Gross_Pay_gas = pay_calculator(Gross_pay_gas_df)
                                well_postcutoff_GOC = well_postcutoff_GOC[well_postcutoff_GOC['Sw']<=Saturation_Cutoff]
                                well_postcutoff_GOC =well_postcutoff_GOC.sort_values(by='MD')
                                Net_Pay_gas = pay_calculator(well_postcutoff_GOC)
                                Porosity_gas = well_postcutoff_GOC['Pi'].mean()
                                Saturation_gas = well_postcutoff_GOC['Sw'].mean()
                                well_postcutoff_OWC = session_state.dict_Rawdataframe[key][(session_state.dict_Rawdataframe[key]['MD'] > GOC) & (session_state.dict_Rawdataframe[key]['MD'] <= OWC)]
                                well_postcutoff_OWC = well_postcutoff_OWC[(well_postcutoff_OWC['Vcl'] <= Vclay_Cutoff) & (well_postcutoff_OWC['Pi'] >= Porosity_Cutoff)]
                                Gross_pay_Oil_df=well_postcutoff_OWC.sort_values(by = 'MD')
                                Gross_Pay_Oil = pay_calculator(Gross_pay_Oil_df)
                                well_postcutoff_OWC = well_postcutoff_OWC[well_postcutoff_OWC['Sw']<=Saturation_Cutoff]
                                well_postcutoff_OWC =well_postcutoff_OWC.sort_values(by='MD')
                                Net_Pay_Oil = pay_calculator(well_postcutoff_OWC)
                                Porosity_Oil = well_postcutoff_OWC['Pi'].mean()
                                Saturation_Oil = well_postcutoff_OWC['Sw'].mean()
                                summary_list = [Gross_Pay_Oil ,Net_Pay_Oil,Porosity_Oil,Saturation_Oil,Gross_Pay_gas,Net_Pay_gas,Porosity_gas,Saturation_gas]
                                summary.extend(summary_list)
                                session_state.dict_Pay_P[key]=summary

                        
                session_state.df_final_case=pd.DataFrame.from_dict(session_state.dict_Pay_P,orient='index')
                session_state.df_final_case.columns= ['Gross_Pay_Oil' ,'Net_Pay_Oil','Porosity_Oil','Saturation_Oil','Gross_Pay_Gas','Net_Pay_Gas','Porosity_Gas','Saturation_Gas']
                Average_Porosity_Oil=((session_state.df_final_case['Net_Pay_Oil']*session_state.df_final_case['Porosity_Oil']).sum())/session_state.df_final_case['Net_Pay_Oil'].sum()
                Average_Saturation_Oil=((session_state.df_final_case['Net_Pay_Oil']*session_state.df_final_case['Porosity_Oil']*session_state.df_final_case['Saturation_Oil']).sum())/(session_state.df_final_case['Net_Pay_Oil']*session_state.df_final_case['Porosity_Oil']).sum()
                Average_Porosity_Gas=((session_state.df_final_case['Net_Pay_Gas']*session_state.df_final_case['Porosity_Gas']).sum())/session_state.df_final_case['Net_Pay_Gas'].sum()
                Average_Saturation_Gas=((session_state.df_final_case['Net_Pay_Gas']*session_state.df_final_case['Porosity_Gas']*session_state.df_final_case['Saturation_Gas']).sum())/(session_state.df_final_case['Net_Pay_Gas']*session_state.df_final_case['Porosity_Gas']).sum()
                summary_list_1 = [Average_Porosity_Oil ,Average_Saturation_Oil,Average_Porosity_Gas,Average_Saturation_Gas]
                summary_Average_property.extend(summary_list_1)
                session_state.cases[f"Vclay {Vclay_Cutoff}, Porosity {Porosity_Cutoff}, Saturation {Saturation_Cutoff}"] = session_state.df_final_case
                session_state.Average_Property[f"Vclay {Vclay_Cutoff}, Porosity {Porosity_Cutoff}, Saturation {Saturation_Cutoff}"] = summary_Average_property
                session_state.Average_Property_df = pd.DataFrame(session_state.Average_Property)
                session_state.Average_Property_df.index = ['Average_Porosity_Oil' ,'Average_Saturation_Oil','Average_Porosity_Gas','Average_Saturation_Gas']
                

        if session_state.cases and len(session_state.cases)==session_state.No_of_rows:
            case = st.selectbox("Chose the case",session_state.cases.keys())
            if case:
                st.write(session_state.cases[case])
                st.write(session_state.Average_Property_df[case].T)  
    with tab2:
        if session_state.cases and len(session_state.cases)==session_state.No_of_rows:
            first_df = next(iter(session_state.cases.values()))
                # Extract the first column of this DataFrame
            first_column = first_df.index
            first_column_list = first_column.tolist()
            keys_list = list(session_state.cases.keys())

                # Initialize an empty list to store the concatenated DataFrames for each column
            concatenated_columns = {}

                # Assuming all DataFrames have the same number of columns
            num_columns = 8

                # Iterate over the column indices
            for col_idx in range(num_columns):
                    # List to store the first column of each DataFrame
                column_list = []
                    
                    # Iterate over the dictionary values (DataFrames)
                for df in session_state.cases.values():
                        # Append the specific column to the list
                    column_list.append(df.iloc[:, col_idx])
                    
                    # Concatenate the columns into a new DataFrame
                concatenated_column_df = pd.concat(column_list, axis=1, ignore_index=True)
                concatenated_column_df.columns = keys_list
                concatenated_column_df.index = first_column_list
                concatenated_columns[session_state.df_final_case.columns[col_idx]]=(concatenated_column_df)

                # Create a select box for user to choose which concatenated DataFrame to plot
            selected_option = st.selectbox("Select a Property to plot:",list(concatenated_columns.keys()) )
                # Plot the selected concatenated DataFrame with Plotly
            selected_df = concatenated_columns[selected_option]
            st.write(selected_df)
            fig = go.Figure()

                # Plot each column
            for column in selected_df.columns:
                fig.add_trace(go.Scatter(x=selected_df.index, y=selected_df[column], mode='lines+markers', name=column))

            fig.update_layout(
                title=f'Line Plot of {selected_option}',
                xaxis_title='Wells',
                yaxis_title='Values',
                legend_title='Cases'
            )
            st.plotly_chart(fig)
            st.success("WEIGHTED AVERAGE PARAMETERS")
            st.write(session_state.Average_Property_df) 
            df_t = session_state.Average_Property_df.T
            fig = go.Figure()
            fig.add_trace(go.Bar(x=df_t.index,y=df_t['Average_Porosity_Oil'],name='Average_Porosity_Oil'))
            fig.add_trace(go.Bar(x=df_t.index,y=df_t['Average_Porosity_Gas'],name='Average_Porosity_Gas'))
            fig.update_layout(title='Bar Plot for Porosity of Different Cases',xaxis_title='Cases',yaxis_title='Values',barmode='group')
            st.plotly_chart(fig) 
        
            fig = go.Figure()
            fig.add_trace(go.Bar(x=df_t.index,y=df_t['Average_Saturation_Oil'],name='Average_Saturation_Oil'))
            fig.add_trace(go.Bar(x=df_t.index,y=df_t['Average_Saturation_Gas'],name='Average_Saturation_Gas'))
            fig.update_layout(title='Bar Plot for Water Saturation of Different Cases',xaxis_title='Cases',yaxis_title='Values',barmode='group')
            st.plotly_chart(fig) 
            
    

# Main function
def main():
    # Create sidebar with radio buttons
    selected_page = st.sidebar.radio('CUTOFF ANALYSIS', ('RAW_DATA_UPLOADER','DEVIATION_DATA_UPLOADER', 'DISPLAYING_THE_RAW DATA', 'VCLAY_ANALYSIS', 'POROSITY_ANALYSIS','SW_ANALYSIS','PAY_AND_PARAMETERS_SUMMARY','SENSITIVITY ANALYSIS'))

    # Conditionally display pages based on selected radio button
    if selected_page == 'RAW_DATA_UPLOADER':
        page0()
    elif selected_page == 'DEVIATION_DATA_UPLOADER':
        page01()
    elif selected_page == 'DISPLAYING_THE_RAW DATA':
        page1()
    elif selected_page == 'VCLAY_ANALYSIS':
        page2()
    elif selected_page == 'POROSITY_ANALYSIS':
        page3()
    elif selected_page == 'SW_ANALYSIS':
        page4()
    elif selected_page == 'PAY_AND_PARAMETERS_SUMMARY':
        page5()
    elif selected_page == 'SENSITIVITY ANALYSIS':
        page6()

if __name__ == '__main__':
    main()
